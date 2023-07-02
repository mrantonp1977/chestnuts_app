import tkinter.ttk
from tkinter import *
from PIL import ImageTk, Image
import sqlite3
import csv
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import re


if __name__ == "__main__":
    root = Tk()
    root.title("ΚΑΣΤΑΝΑ")
    root.geometry("710x733")
    root.config(background="#EEE8D1")
    icon = PhotoImage(file="chestnuts2.png")
    root.iconphoto(True, icon)




    conn = sqlite3.connect("kastana_file.db")
    c = conn.cursor()


    c.execute("""CREATE TABLE IF NOT EXISTS products (
    date integer,
    extra_size_kilos integer,
    extra_size_price integer,
    a_size_kilos integer,
    a_size_price integer,
    b_size_kilos integer,
    b_size_price integer,
    c_size_kilos integer,
    c_size_price integer,
    torn_kilos integer,
    torn_price integer)""")



    title_label = Label(root, text="   ΚΑΣΤΑΝΑ DATABASE   ", font=("arial black", 20), background="black",
                        relief="ridge", borderwidth=7, foreground="#F60505")
    title_label.grid(row=0, column=0, columnspan=3, padx=20, pady=15)




    def update():
        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()
        show_id = delete_box.get()
        c.execute("""UPDATE products SET
               date = :date,
               extra_size_kilos = :extra_size_kilos,
               extra_size_price = :extra_size_price,
               a_size_kilos = :a_size_kilos,
               a_size_price = :a_size_price,
               b_size_kilos = :b_size_kilos,
               b_size_price = :b_size_price,
               c_size_kilos = :c_size_kilos,
               c_size_price = :c_size_price,
               torn_kilos = :torn_kilos,
               torn_price = :torn_price
    
    
               WHERE oid = :oid""",
                  {'date': date_editor.get(),
                   'extra_size_kilos': extra_size_kilos_editor.get(),
                   'extra_size_price': extra_size_price_editor.get(),
                   'a_size_kilos': a_size_kilos_editor.get(),
                   'a_size_price': a_size_price_editor.get(),
                   'b_size_kilos': b_size_kilos_editor.get(),
                   'b_size_price': b_size_price_editor.get(),
                   'c_size_kilos': c_size_kilos_editor.get(),
                   'c_size_price': c_size_price_editor.get(),
                   'torn_kilos': torn_kilos_editor.get(),
                   'torn_price': torn_price_editor.get(),
                   'oid': show_id})

        confirmation = messagebox.askquestion("Επεξεργασία Αρχείων", "Do you want to save the changes  ???")
        if confirmation == 'yes':
            conn.commit()
            messagebox.showinfo(" Αποθήκευση Αρχείων ", "Επιτυχής αλλαγή και αποθήκευση Αρχείων. !!!")
        else:
            conn.rollback()
            messagebox.showinfo("Ακύρωση Εισαγωγής", "Η αλλαγή Αρχείων ακυρώθηκε. !!!")

        conn.close()
        editor.destroy()

    def save_to_excel():
        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()

        c.execute("SELECT * FROM products")
        data = c.fetchall()

        try:
            workbook = load_workbook("total.chestnuts.xlsx")
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            headers = ["ΗΜΕΡΟΜΗΝΙΑ", " EXTRA ", " ΤΙΜΗ ", " Α ", " ΤΙΜΗ ", " Β ", " ΤΙΜΗ ", " Γ ", " ΤΙΜΗ ", " ΣΚΙΣΜΕΝΑ ", " ΤΙΜΗ "]
            sheet.append(headers)

        sheet.delete_rows(2, sheet.max_row)

        for row in data:
            sheet.append(row)

        # Save the workbook to a file
        workbook.save("total.chestnuts.xlsx")

        messagebox.showinfo("Αποθήκευση Δεδομένων", "Τα δεδομένα αποθηκεύτηκαν σε αρχείο Excel !!!")

        conn.close()

    def submit():
        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()

        c.execute("INSERT INTO products VALUES (:date, :extra_size_kilos, :extra_size_price, :a_size_kilos, :a_size_price, :b_size_kilos, :b_size_price, :c_size_kilos, :c_size_price, :torn_kilos, :torn_price)",
                  {
                      "date": date.get(),
                      "extra_size_kilos": extra_size_kilos.get() + " ΚΙΛΑ",
                      "extra_size_price": extra_size_price.get() + " €",
                      "a_size_kilos": a_size_kilos.get() + " ΚΙΛΑ",
                      "a_size_price": a_size_price.get() + " €",
                      "b_size_kilos": b_size_kilos.get() + " ΚΙΛΑ",
                      "b_size_price": b_size_price.get() + " €",
                      "c_size_kilos": c_size_kilos.get() + " ΚΙΛΑ",
                      "c_size_price": c_size_price.get() + " €",
                      "torn_kilos": torn_kilos.get() + " ΚΙΛΑ",
                      "torn_price": torn_price.get() + " €"
                  })

        confirmation = messagebox.askquestion("Εισαγωγή Αρχείων", "Do You Want To Insert this Files ???")
        if confirmation == 'yes':
            conn.commit()
            messagebox.showinfo(" Εισαγωγή Αρχείων ", "Επιτυχής Εισαγωγή Αρχείων. !!!")
        else:
            conn.rollback()
            messagebox.showinfo("Ακύρωση Εισαγωγής", "Η Εισαγωγή Αρχείων ακυρώθηκε. !!!")
        conn.close()

        date.delete(0, END)
        extra_size_kilos.delete(0, END)
        extra_size_price.delete(0, END)
        a_size_kilos.delete(0, END)
        a_size_price.delete(0, END)
        b_size_kilos.delete(0, END)
        b_size_price.delete(0, END)
        c_size_kilos.delete(0, END)
        c_size_price.delete(0, END)
        torn_kilos.delete(0, END)
        torn_price.delete(0, END)

    def delete_all_data():
        confirmation = messagebox.askyesno("Confirmation", "Είστε σίγουρος ότι θέλετε να διαγράψετε όλα τα Αρχεία ;")
        if confirmation:
            conn = sqlite3.connect("kastana_file.db")
            c = conn.cursor()
            c.execute("DELETE FROM products")
            conn.commit()
            conn.close()
            messagebox.showinfo("Success", "Ολα τα Αρχεία διαγράφηκαν με επιτυχία !!")



    def show():
        show = Tk()
        show.title("ΚΑΣΤΑΝΑ DATABASE")
        show.geometry("990x800")
        show.config(background="#E9E8D4")

        show_id = delete_box.get()
        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()

        c.execute("SELECT *,  oid FROM products")
        elements = c.fetchall()

        for i, element in enumerate(elements):
            text = '    '.join(str(item) for item in element)
            show_label = Label(show, text=text, font=("arial black", 11), relief="ridge", bd=10, borderwidth=5,
                               background="#E9E8D4", foreground="#20047A")
            show_label.grid(row=i, column=0, pady=5, sticky=W, ipadx=30)

        save_button = Button(show, text="Αποθήκευση σε 'Excel' ", command=save_to_excel, font=("arial black", 11),
                             background="#06D784")
        save_button.grid(row=i + 1, column=0, pady=5, padx=120, sticky=W,  ipadx=20)

        delete_button = Button(show, text="Διαγραφή όλων των Αρχείων", command=delete_all_data, font=("arial black", 11),
                               background="#EA6969")
        delete_button.grid(row=i + 1, column=0, pady=10, padx=550, ipadx=8)

    def delete():
        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()

        c.execute("DELETE from products WHERE oid = " + delete_box.get())

        confirmation = messagebox.askquestion("Διαγραφή Στοιχείων", "Θέλετε να διαγράψετε τα Στοιχεία ;")
        if confirmation == 'yes':
            conn.commit()
            messagebox.showinfo("Η Διαγραφή Ολοκληρώθηκε", "Τα στοιχεία διαγράφηκαν με επιτυχία. !!!")
        else:
            conn.rollback()
            messagebox.showinfo("Ακύρωση Διαγραφής", "Η διαγραφή ακυρώθηκε. !!!")

        conn.commit()
        conn.close()

    def edit():
        global editor
        editor = Tk()
        editor.title("ΕΠΕΞΕΡΓΑΣΙΑ ΣΤΟΙΧΕΙΩΝ")
        editor.geometry("850x550")
        editor.config(background="#EEE8D1")


        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()

        show_id = delete_box.get()

        c.execute("SELECT * FROM products WHERE oid = " + show_id)
        shows = c.fetchall()

        global date_editor
        global extra_size_kilos_editor
        global extra_size_price_editor
        global a_size_kilos_editor
        global a_size_price_editor
        global b_size_kilos_editor
        global b_size_price_editor
        global c_size_kilos_editor
        global c_size_price_editor
        global torn_kilos_editor
        global torn_price_editor

        date_editor = Label(editor, text="Ημερομηνία", font=("arial black", 16), background="#EEE8D1", foreground="#5D046B")
        date_editor.grid(row=1, column=0, pady=(20, 0), sticky=W, padx=80)
        extra_size_kilos_editor = Label(editor, text="ΜΕΓΕΘΟΣ 'EXTRA'", font=("arial black", 14), foreground="#5D046B",
                                       background="#EEE8D1")
        extra_size_kilos_editor.grid(row=3, column=0, sticky=W, padx=10, pady=10)
        a_size_kilos_editor = Label(editor, text="ΜΕΓΕΘΟΣ  'Α'", font=("arial black", 14), foreground="#5D046B",
                                   background="#EEE8D1")
        a_size_kilos_editor.grid(row=4, column=0, sticky=W, padx=10, pady=10)
        b_size_kilos_editor = Label(editor, text="ΜΕΓΕΘΟΣ  'Β'", font=("arial black", 14), foreground="#5D046B",
                                   background="#EEE8D1")
        b_size_kilos_editor.grid(row=5, column=0, sticky=W, padx=10, pady=10)
        c_size_kilos_editor = Label(editor, text="ΜΕΓΕΘΟΣ  'Γ'", font=("arial black", 14), foreground="#5D046B",
                                   background="#EEE8D1")
        c_size_kilos_editor.grid(row=6, column=0, sticky=W, padx=10, pady=10)
        torn_kilos_editor = Label(editor, text="ΣΚΙΣΜΕΝΑ", font=("arial black", 14), foreground="#5D046B",
                                 background="#EEE8D1")
        torn_kilos_editor.grid(row=7, column=0, sticky=W, padx=10, pady=10)

        date_editor = Entry(editor, width=12, font=("arial black", 14), bd=2)
        date_editor.grid(row=1, column=1, columnspan=2, padx=50, pady=(20, 5))
        extra_size_kilos_editor = Entry(editor, width=12, font=("arial black", 14), bd=2)
        extra_size_kilos_editor.grid(row=3, column=1, padx=50, pady=10)
        extra_size_price_editor = Entry(editor, width=12, font=("arial black", 14), bd=2)
        extra_size_price_editor.grid(row=3, column=2, padx=50, pady=10)
        a_size_kilos_editor = Entry(editor, width=12, font=("arial black", 14), bd=2)
        a_size_kilos_editor.grid(row=4, column=1, padx=50, pady=10)
        a_size_price_editor = Entry(editor, width=12, font=("arial black", 14), bd=2)
        a_size_price_editor.grid(row=4, column=2, padx=50, pady=10)
        b_size_kilos_editor = Entry(editor, width=12, font=("arial black", 14), bd=2)
        b_size_kilos_editor.grid(row=5, column=1, padx=50, pady=10)
        b_size_price_editor = Entry(editor, width=12, font=("arial black", 14), bd=2)
        b_size_price_editor.grid(row=5, column=2, padx=50, pady=10)
        c_size_kilos_editor = Entry(editor, width=12, font=("arial black", 14), bd=2)
        c_size_kilos_editor.grid(row=6, column=1, padx=50, pady=10)
        c_size_price_editor = Entry(editor, width=12, font=("arial black", 14), bd=2)
        c_size_price_editor.grid(row=6, column=2, padx=50, pady=10)
        torn_kilos_editor = Entry(editor, width=12, font=("arial black", 14), bd=2)
        torn_kilos_editor.grid(row=7, column=1, padx=50, pady=10)
        torn_price_editor = Entry(editor, width=12, font=("arial black", 14), bd=2)
        torn_price_editor.grid(row=7, column=2, padx=50, pady=10)

        for show in shows:
            date_editor.insert(0, show[0])
            extra_size_kilos_editor.insert(0, show[1])
            extra_size_price_editor.insert(0, show[2])
            a_size_kilos_editor.insert(0, show[3])
            a_size_price_editor.insert(0, show[4])
            b_size_kilos_editor.insert(0, show[5])
            b_size_price_editor.insert(0, show[6])
            c_size_kilos_editor.insert(0, show[7])
            c_size_price_editor.insert(0, show[8])
            torn_kilos_editor.insert(0, show[9])
            torn_price_editor.insert(0, show[10])

        edit_btn = Button(editor, text="Επεξεργασία και Αποθήκευση \n Στοιχείων", bd=6, command=update,
                          font=("arial black", 11), background="#1CC405")
        edit_btn.grid(row=8, column=0, columnspan=3, pady=30, padx=10, ipadx=30)



    total_extra_size_kilos_label = None
    total_a_size_kilos_label = None
    total_b_size_kilos_label = None
    total_c_size_kilos_label = None
    total_torn_kilos_label = None
    total_all_kilos_label = None

    total_extra_size_price_label = None
    total_a_price_label = None
    total_b_price_label = None
    total_c_price_label = None
    total_torn_price_label = None
    total_average_label = None

    total_extra_size_euros_label = None
    total_a_size_euros_label = None
    total_b_size_euros_label = None
    total_c_size_euros_label = None
    total_torn_euros_label = None
    total_all_euros_label = None





    def total_extra_size_kilos():
        global total_extra_size_kilos_label
        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()
        c.execute("SELECT SUM(extra_size_kilos) FROM products")
        result = c.fetchone()[0]
        total_extra_size_kilos_value = 0 if result is None else int(result)

        total_extra_size_kilos_label.config(text=" 'EXTRA' ΚΙΛΑ  : {} ".format(total_extra_size_kilos_value))

    def total_a_size_kilos():
        global total_a_size_kilos_label
        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()
        c.execute("SELECT SUM(a_size_kilos) FROM products")
        result = c.fetchone()[0]
        total_a_size_kilos_value = 0 if result is None else int(result)

        total_a_size_kilos_label.config(text=" 'A' ΚΙΛΑ  : {} ".format(total_a_size_kilos_value))

    def total_b_size_kilos():
        global total_b_size_kilos_label
        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()
        c.execute("SELECT SUM(b_size_kilos) FROM products")
        result = c.fetchone()[0]
        total_b_size_kilos_value = 0 if result is None else int(result)

        total_b_size_kilos_label.config(text=" 'B' ΚΙΛΑ  : {} ".format(total_b_size_kilos_value))


    def total_c_size_kilos():
        global total_c_size_kilos_label
        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()
        c.execute("SELECT SUM(c_size_kilos) FROM products")
        result = c.fetchone()[0]
        total_c_size_kilos_value = 0 if result is None else int(result)

        total_c_size_kilos_label.config(text=" 'Γ' ΚΙΛΑ  : {} ".format(total_c_size_kilos_value))


    def total_torn_kilos():
        global total_torn_kilos_label
        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()
        c.execute("SELECT SUM(torn_kilos) FROM products")
        result = c.fetchone()[0]
        total_torn_kilos_value = 0 if result is None else int(result)

        total_torn_kilos_label.config(text=" 'ΣΚΙΣΜΕΝΑ' ΚΙΛΑ  : {} ".format(total_torn_kilos_value))




    def total_all_kilos():
        global total_all_kilos_label

        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()

        c.execute(
            "SELECT SUM(extra_size_kilos + a_size_kilos + b_size_kilos + c_size_kilos + torn_kilos) FROM products")
        result = c.fetchone()[0]
        total_all_kilos_value = 0 if result is None else int(result)

        total_all_kilos_label.config(text="ΣΥΝΟΛΙΚΑ  ΚΙΛΑ  : {} ".format(total_all_kilos_value))



    def open_kilos_w():
        global total_extra_size_kilos_label
        global total_a_size_kilos_label
        global total_b_size_kilos_label
        global total_c_size_kilos_label
        global total_torn_kilos_label
        global total_all_kilos_label

        kilos_w = Tk()
        kilos_w.title("ΣΥΝΟΛΙΚΑ ΚΙΛΑ")
        kilos_w.geometry("800x600")
        kilos_w.config(background="#EEE8D1")


        total_extra_size_kilos_btn = Button(kilos_w, text="ΣΥΝΟΛΙΚΑ ΚΙΛΑ 'EXTRA'", command=total_extra_size_kilos, font=("arial black", 11), bd=4, background="#DE9C30", activeforeground="#DE9C30", activebackground="#DE9C30")
        total_extra_size_kilos_btn.grid(row=1, column=0, pady=(25, 5), padx=10, ipadx=57)
        total_extra_size_kilos_label = Label(kilos_w, text="   ", font=("arial black", 13), background="#A2D8F0",
                             relief="ridge", borderwidth=5, foreground="#42047A", width=15)
        total_extra_size_kilos_label.grid(row=1, column=1, ipadx=100, pady=(25, 5))

        total_a_size_kilos_btn = Button(kilos_w, text="ΣΥΝΟΛΙΚΑ ΚΙΛΑ 'Α'", command=total_a_size_kilos, font=("arial black", 11), bd=4, background="#DE9C30", activeforeground="#DE9C30", activebackground="#DE9C30")
        total_a_size_kilos_btn.grid(row=2, column=0, padx=10, ipadx=79, pady=5)
        total_a_size_kilos_label = Label(kilos_w, text="   ", font=("arial black", 13), background="#A2D8F0",
                             relief="ridge", borderwidth=5, foreground="#42047A", width=15)
        total_a_size_kilos_label.grid(row=2, column=1, ipadx=100, pady=5)

        total_b_size_kilos_btn = Button(kilos_w, text="ΣΥΝΟΛΙΚΑ ΚΙΛΑ 'Β'", command=total_b_size_kilos, font=("arial black", 11), bd=4, background="#DE9C30", activeforeground="#DE9C30", activebackground="#DE9C30")
        total_b_size_kilos_btn.grid(row=3, column=0, padx=10, ipadx=79, pady=5)
        total_b_size_kilos_label = Label(kilos_w, text="   ", font=("arial black", 13), background="#A2D8F0",
                             relief="ridge", borderwidth=5, foreground="#42047A", width=15)
        total_b_size_kilos_label.grid(row=3, column=1, ipadx=100, pady=5)

        total_c_size_kilos_btn = Button(kilos_w, text="ΣΥΝΟΛΙΚΑ ΚΙΛΑ 'Γ'", command=total_c_size_kilos, font=("arial black", 11), bd=4, background="#DE9C30", activeforeground="#DE9C30", activebackground="#DE9C30")
        total_c_size_kilos_btn.grid(row=4, column=0, padx=10, ipadx=79, pady=5)
        total_c_size_kilos_label = Label(kilos_w, text="   ", font=("arial black", 13), background="#A2D8F0",
                             relief="ridge", borderwidth=5, foreground="#42047A", width=15)
        total_c_size_kilos_label.grid(row=4, column=1, ipadx=100, pady=5)

        total_torn_kilos_btn = Button(kilos_w, text="ΣΥΝΟΛΙΚΑ ΚΙΛΑ 'ΣΚΙΣΜΕΝΑ'", command=total_torn_kilos, font=("arial black", 11), bd=4, background="#DE9C30", activeforeground="#DE9C30", activebackground="#DE9C30")
        total_torn_kilos_btn.grid(row=5, column=0, padx=10, ipadx=40, pady=5)
        total_torn_kilos_label = Label(kilos_w, text="   ", font=("arial black", 13), background="#A2D8F0",
                             relief="ridge", borderwidth=5, foreground="#42047A", width=15)
        total_torn_kilos_label.grid(row=5, column=1, ipadx=100, pady=5)

        total_all_kilos_btn = Button(kilos_w, text="ΣΥΝΟΛΟ ΚΙΛΩΝ", command=total_all_kilos, font=("arial black", 13), bd=7, background="#1CC405", activeforeground="#1CC405", activebackground="#1CC405")
        total_all_kilos_btn.grid(row=6, column=0, columnspan=2,  padx=10, ipadx=79, pady=45)
        total_all_kilos_label = Label(kilos_w, text="     ", font=("arial black", 15), background="#D1EFCD",
                             relief="ridge", borderwidth=10, foreground="#42047A", width=20)
        total_all_kilos_label.grid(row=7, column=0,columnspan=2, ipadx=100, pady=5)

        title_label = Label(kilos_w, text=" ΚΑΣΤΑΝΑ DATABASE ", font=("arial black", 18), background="black",
                            relief="ridge", borderwidth=10, foreground="red")
        title_label.grid(row=0, pady=10, padx=10, columnspan=2)



    def total_extra_size_price():
        global total_extra_size_price_label

        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()

        c.execute("SELECT extra_size_price FROM products")
        prices = c.fetchall()

        valid_prices = []
        for price in prices:
            price_str = str(price[0])
            price_str = re.sub(r'[^0-9.,]', '', price_str)
            price_str = price_str.replace(',', '.')
            if price_str and float(price_str) != 0:
                valid_prices.append(float(price_str))

        total_extra_size_price = sum(valid_prices)
        average_price = total_extra_size_price / len(valid_prices) if valid_prices else 0

        num_valid_prices = len(valid_prices)
        if num_valid_prices > 0:
            average_price = total_extra_size_price / num_valid_prices
        else:
            average_price = 0

        total_extra_size_price_label.config(text="ΜΕΣΗ ΤΙΜΗ: {:.2f} €".format(average_price))




    def total_a_size_price():
        global total_a_size_price_label

        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()

        c.execute("SELECT a_size_price FROM products")
        prices = c.fetchall()

        valid_prices = []
        for price in prices:
            price_str = str(price[0])
            price_str = re.sub(r'[^0-9.,]', '', price_str)
            price_str = price_str.replace(',', '.')
            if price_str and float(price_str) != 0:
                valid_prices.append(float(price_str))

        total_a_size_price = sum(valid_prices)
        average_price = total_a_size_price / len(valid_prices) if valid_prices else 0

        num_valid_prices = len(valid_prices)
        if num_valid_prices > 0:
            average_price = total_a_size_price / num_valid_prices
        else:
            average_price = 0

        total_a_size_price_label.config(text="ΜΕΣΗ ΤΙΜΗ: {:.2f} €".format(average_price))



    def total_b_size_price():
        global total_b_size_price_label

        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()

        c.execute("SELECT b_size_price FROM products")
        prices = c.fetchall()

        valid_prices = []
        for price in prices:
            price_str = str(price[0])
            price_str = re.sub(r'[^0-9.,]', '', price_str)
            price_str = price_str.replace(',', '.')
            if price_str and float(price_str) != 0:
                valid_prices.append(float(price_str))

        total_b_size_price = sum(valid_prices)
        average_price = total_b_size_price / len(valid_prices) if valid_prices else 0

        num_valid_prices = len(valid_prices)
        if num_valid_prices > 0:
            average_price = total_b_size_price / num_valid_prices
        else:
            average_price = 0

        total_b_size_price_label.config(text="ΜΕΣΗ ΤΙΜΗ: {:.2f} €".format(average_price))

    def total_c_size_price():
        global total_c_size_price_label

        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()

        c.execute("SELECT c_size_price FROM products")
        prices = c.fetchall()

        valid_prices = []
        for price in prices:
            price_str = str(price[0])
            price_str = re.sub(r'[^0-9.,]', '', price_str)
            price_str = price_str.replace(',', '.')
            if price_str and float(price_str) != 0:
                valid_prices.append(float(price_str))

        total_c_size_price = sum(valid_prices)
        average_price = total_c_size_price / len(valid_prices) if valid_prices else 0

        num_valid_prices = len(valid_prices)
        if num_valid_prices > 0:
            average_price = total_c_size_price / num_valid_prices
        else:
            average_price = 0

        total_c_size_price_label.config(text="ΜΕΣΗ ΤΙΜΗ: {:.2f} €".format(average_price))

    def total_torn_price():
        global total_torn_price_label

        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()

        c.execute("SELECT torn_price FROM products")
        prices = c.fetchall()

        valid_prices = []
        for price in prices:
            price_str = str(price[0])
            price_str = re.sub(r'[^0-9.,]', '', price_str)
            price_str = price_str.replace(',', '.')
            if price_str and float(price_str) != 0:
                valid_prices.append(float(price_str))

        total_torn_price = sum(valid_prices)
        average_price = total_torn_price / len(valid_prices) if valid_prices else 0

        num_valid_prices = len(valid_prices)
        if num_valid_prices > 0:
            average_price = total_torn_price / num_valid_prices
        else:
            average_price = 0

        total_torn_price_label.config(text="ΜΕΣΗ ΤΙΜΗ: {:.2f} €".format(average_price))


    def total_average():
        global total_all_euros_label
        global total_all_kilos_label

        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()

        c.execute(
            "SELECT extra_size_kilos, extra_size_price, a_size_kilos, a_size_price, b_size_kilos, b_size_price, c_size_kilos, c_size_price, torn_kilos, torn_price FROM products")
        rows = c.fetchall()

        total_all_kilos = 0
        total_all_euros = 0

        for row in rows:
            extra_kilos = row[0]
            extra_price = row[1]
            a_kilos = row[2]
            a_price = row[3]
            b_kilos = row[4]
            b_price = row[5]
            c_kilos = row[6]
            c_price = row[7]
            torn_kilos = row[8]
            torn_price = row[9]

            if extra_kilos and extra_price and extra_kilos.strip() and extra_kilos.strip() != 'ΚΙΛΑ':
                extra_kilos = float(''.join(filter(str.isdigit, extra_kilos)))
                if extra_price.strip() and extra_price.strip() != '€':
                    extra_price = float(str(extra_price).replace(',', '.').replace('€', ''))
                    total_all_kilos += extra_kilos
                    total_all_euros += extra_kilos * extra_price

            if a_kilos and a_price and a_kilos.strip() and a_kilos.strip() != 'ΚΙΛΑ':
                a_kilos = float(''.join(filter(str.isdigit, a_kilos)))
                if a_price.strip() and a_price.strip() != '€':
                    a_price = float(str(a_price).replace(',', '.').replace('€', ''))
                    total_all_kilos += a_kilos
                    total_all_euros += a_kilos * a_price

            if b_kilos and b_price and b_kilos.strip() and b_kilos.strip() != 'ΚΙΛΑ':
                b_kilos = float(''.join(filter(str.isdigit, b_kilos)))
                if b_price.strip() and b_price.strip() != '€':
                    b_price = float(str(b_price).replace(',', '.').replace('€', ''))
                    total_all_kilos += b_kilos
                    total_all_euros += b_kilos * b_price

            if c_kilos and c_price and c_kilos.strip() and c_kilos.strip() != 'ΚΙΛΑ':
                c_kilos = float(''.join(filter(str.isdigit, c_kilos)))
                if c_price.strip() and c_price.strip() != '€':
                    c_price = float(str(c_price).replace(',', '.').replace('€', ''))
                    total_all_kilos += c_kilos
                    total_all_euros += c_kilos * c_price

            if torn_kilos and torn_price and torn_kilos.strip() and torn_kilos.strip() != 'ΚΙΛΑ':
                torn_kilos = float(''.join(filter(str.isdigit, torn_kilos)))
                if torn_price.strip() and torn_price.strip() != '€':
                    torn_price = float(str(torn_price).replace(',', '.').replace('€', ''))
                    total_all_kilos += torn_kilos
                    total_all_euros += torn_kilos * torn_price

        conn.close()

        if total_all_kilos != 0:
            average_price = total_all_euros / total_all_kilos
        else:
            average_price = 0.0

        total_average_label.config(text="ΜΕΣΗ ΤΙΜΗ : {:.2f} €".format(average_price))

        return average_price


    def average_price():
        global total_extra_size_price_label
        global total_a_size_price_label
        global total_b_size_price_label
        global total_c_size_price_label
        global total_torn_price_label
        global total_average_label

        price_w = Tk()
        price_w.title("ΤΙΜΕΣ")
        price_w.geometry("800x600")
        price_w.config(background="#EEE8D1")



        total_extra_size_price_btn = Button(price_w, text="ΜΕΣΗ ΤΙΜΗ 'EXTRA'", command=total_extra_size_price, font=("arial black", 11), bd=4, background="#DE9C30", activeforeground="#DE9C30", activebackground="#DE9C30")
        total_extra_size_price_btn.grid(row=1, column=0, pady=(25, 5), padx=10, ipadx=57)
        total_extra_size_price_label = Label(price_w, text="   ", font=("arial black", 14), background="#A2D8F0",
                             relief="ridge", borderwidth=5, foreground="#42047A", width=15)
        total_extra_size_price_label.grid(row=1, column=1, ipadx=100, pady=(25, 5))

        total_a_size_price_btn = Button(price_w, text="ΜΕΣΗ ΤΙΜΗ 'Α'", command=total_a_size_price, font=("arial black", 11), bd=4, background="#DE9C30", activeforeground="#DE9C30", activebackground="#DE9C30")
        total_a_size_price_btn.grid(row=2, column=0, pady=5, padx=10, ipadx=80)
        total_a_size_price_label = Label(price_w, text="   ", font=("arial black", 14), background="#A2D8F0",
                             relief="ridge", borderwidth=5, foreground="#42047A", width=15)
        total_a_size_price_label.grid(row=2, column=1, ipadx=100, pady=5)

        total_b_size_price_btn = Button(price_w, text="ΜΕΣΗ ΤΙΜΗ 'Β'", command=total_b_size_price, font=("arial black", 11), bd=4, background="#DE9C30", activeforeground="#DE9C30", activebackground="#DE9C30")
        total_b_size_price_btn.grid(row=3, column=0, pady=5, padx=10, ipadx=80)
        total_b_size_price_label = Label(price_w, text="   ", font=("arial black", 14), background="#A2D8F0",
                             relief="ridge", borderwidth=5, foreground="#42047A", width=15)
        total_b_size_price_label.grid(row=3, column=1, ipadx=100, pady=5)

        total_c_size_price_btn = Button(price_w, text="ΜΕΣΗ ΤΙΜΗ 'Γ'", command=total_c_size_price, font=("arial black", 11), bd=4, background="#DE9C30", activeforeground="#DE9C30", activebackground="#DE9C30")
        total_c_size_price_btn.grid(row=4, column=0, pady=5, padx=10, ipadx=80)
        total_c_size_price_label = Label(price_w, text="   ", font=("arial black", 14), background="#A2D8F0",
                             relief="ridge", borderwidth=5, foreground="#42047A", width=15)
        total_c_size_price_label.grid(row=4, column=1, ipadx=100, pady=5)

        total_torn_price_btn = Button(price_w, text="ΜΕΣΗ ΤΙΜΗ 'ΣΚΙΣΜΕΝΑ'", command=total_torn_price, font=("arial black", 11), bd=4, background="#DE9C30", activeforeground="#DE9C30", activebackground="#DE9C30")
        total_torn_price_btn.grid(row=5, column=0, pady=5, padx=10, ipadx=39)
        total_torn_price_label = Label(price_w, text="   ", font=("arial black", 14), background="#A2D8F0",
                             relief="ridge", borderwidth=5, foreground="#42047A", width=15)
        total_torn_price_label.grid(row=5, column=1, ipadx=100, pady=5)

        total_average_btn = Button(price_w, text="ΣΥΝΟΛΙΚΗ ΜΕΣΗ ΤΙΜΗ", command=total_average, font=("arial black", 13), bd=8, background="#1CC405", activeforeground="#1CC405", activebackground="#1CC405")
        total_average_btn.grid(row=6, column=0, columnspan=2,  pady=40, padx=10, ipadx=80)
        total_average_label = Label(price_w, text="     ", font=("arial black", 15), background="#D1EFCD",
                             relief="ridge", borderwidth=10, foreground="#42047A", width=20)
        total_average_label.grid(row=7, column=0, columnspan=2, ipadx=100, pady=15)

        title_label = Label(price_w, text=" ΚΑΣΤΑΝΑ DATABASE ", font=("arial black", 18), background="black",
                            relief="ridge", borderwidth=10, foreground="red")
        title_label.grid(row=0, pady=10, padx=10, columnspan=2)

    def total_extra_size_euros():
        global total_extra_size_euros_label
        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()

        c.execute("SELECT extra_size_kilos, extra_size_price FROM products")
        rows = c.fetchall()
        total_extra_size_euros = 0

        for row in rows:
            kilos = row[0]
            price = row[1]
            if kilos and price and kilos.strip() and kilos.strip() != 'ΚΙΛΑ':
                kilos = float(''.join(filter(str.isdigit, kilos)))
                if price.strip() and price.strip() != '€':
                    price = float(str(price).replace(',', '.').replace('€', ''))
                    total_extra_size_euros += kilos * price

        conn.close()

        total_extra_size_euros_label.config(text="'EXTRA' ΕΥΡΩ: {:.2f} €".format(total_extra_size_euros))
        return total_extra_size_euros




    def total_a_size_euros():
        global total_a_size_euros_label
        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()

        c.execute("SELECT a_size_kilos, a_size_price FROM products")
        rows = c.fetchall()
        total_a_size_euros = 0

        for row in rows:
            kilos = row[0]
            price = row[1]
            if kilos and price and kilos.strip() and kilos.strip() != 'ΚΙΛΑ':
                kilos = float(''.join(filter(str.isdigit, kilos)))
                price = float(str(price).replace(',', '.').replace('€', ''))
                total_a_size_euros += kilos * price

        conn.close()

        total_a_size_euros_label.config(text="'Α' ΕΥΡΩ: {:.2f} €".format(total_a_size_euros))
        return total_a_size_euros

    def total_b_size_euros():
        global total_b_size_euros_label
        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()

        c.execute("SELECT b_size_kilos, b_size_price FROM products")
        rows = c.fetchall()
        total_b_size_euros = 0

        for row in rows:
            kilos = row[0]
            price = row[1]
            if kilos and price and kilos.strip() and kilos.strip() != 'ΚΙΛΑ':
                kilos = float(''.join(filter(str.isdigit, kilos)))
                price = float(str(price).replace(',', '.').replace('€', ''))
                total_b_size_euros += kilos * price

        conn.close()

        total_b_size_euros_label.config(text="'Β' ΕΥΡΩ: {:.2f} €".format(total_b_size_euros))
        return total_b_size_euros

    def total_c_size_euros():
        global total_c_size_euros_label
        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()

        c.execute("SELECT c_size_kilos, c_size_price FROM products")
        rows = c.fetchall()
        total_c_size_euros = 0

        for row in rows:
            kilos = row[0]
            price = row[1]
            if kilos and price and kilos.strip() and kilos.strip() != 'ΚΙΛΑ':
                kilos = float(''.join(filter(str.isdigit, kilos)))
                if price.strip() and price.strip() != '€':
                    price = float(str(price).replace(',', '.').replace('€', ''))
                    total_c_size_euros += kilos * price

        conn.close()

        total_c_size_euros_label.config(text="'Γ' ΕΥΡΩ: {:.2f} €".format(total_c_size_euros))
        return total_c_size_euros


    def total_torn_euros():
        global total_torn_euros_label
        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()

        c.execute("SELECT torn_kilos, torn_price FROM products")
        rows = c.fetchall()
        total_torn_euros = 0

        for row in rows:
            kilos = row[0]
            price = row[1]
            if kilos and price and kilos.strip() and kilos.strip() != 'ΚΙΛΑ':
                kilos = float(''.join(filter(str.isdigit, kilos)))
                price = float(str(price).replace(',', '.').replace('€', ''))
                total_torn_euros += kilos * price

        conn.close()

        total_torn_euros_label.config(text="'ΣΚΙΣΜΕΝΑ' ΕΥΡΩ: {:.2f} €".format(total_torn_euros))
        return total_torn_euros

    def total_all_euros():
        global total_all_euros_label
        conn = sqlite3.connect("kastana_file.db")
        c = conn.cursor()

        c.execute("SELECT extra_size_kilos, extra_size_price, a_size_kilos, a_size_price, b_size_kilos, b_size_price, c_size_kilos, c_size_price, torn_kilos, torn_price FROM products")
        rows = c.fetchall()
        total_all_euros = 0

        for row in rows:
            extra_kilos = row[0]
            extra_price = row[1]
            a_kilos = row[2]
            a_price = row[3]
            b_kilos = row[4]
            b_price = row[5]
            c_kilos = row[6]
            c_price = row[7]
            torn_kilos = row[8]
            torn_price = row[9]

            if extra_kilos and extra_price and extra_kilos.strip() and extra_kilos.strip() != 'ΚΙΛΑ':
                extra_kilos = float(''.join(filter(str.isdigit, extra_kilos)))
                if extra_price.strip() and extra_price.strip() != '€':
                    extra_price = float(str(extra_price).replace(',', '.').replace('€', ''))
                    total_all_euros += extra_kilos * extra_price

            if a_kilos and a_price and a_kilos.strip() and a_kilos.strip() != 'ΚΙΛΑ':
                a_kilos = float(''.join(filter(str.isdigit, a_kilos)))
                if a_price.strip() and a_price.strip() != '€':
                    a_price = float(str(a_price).replace(',', '.').replace('€', ''))
                    total_all_euros += a_kilos * a_price

            if b_kilos and b_price and b_kilos.strip() and b_kilos.strip() != 'ΚΙΛΑ':
                b_kilos = float(''.join(filter(str.isdigit, b_kilos)))
                if b_price.strip() and b_price.strip() != '€':
                    b_price = float(str(b_price).replace(',', '.').replace('€', ''))
                    total_all_euros += b_kilos * b_price

            if c_kilos and c_price and c_kilos.strip() and c_kilos.strip() != 'ΚΙΛΑ':
                c_kilos = float(''.join(filter(str.isdigit, c_kilos)))
                if c_price.strip() and c_price.strip() != '€':
                    c_price = float(str(c_price).replace(',', '.').replace('€', ''))
                    total_all_euros += c_kilos * c_price

            if torn_kilos and torn_price and torn_kilos.strip() and torn_kilos.strip() != 'ΚΙΛΑ':
                torn_kilos = float(''.join(filter(str.isdigit, torn_kilos)))
                if torn_price.strip() and torn_price.strip() != '€':
                    torn_price = float(str(torn_price).replace(',', '.').replace('€', ''))
                    total_all_euros += torn_kilos * torn_price

        conn.close()

        total_all_euros_label.config(text="ΣΥΝΟΛΟ ΕΥΡΩ : {:.2f} €".format(total_all_euros))
        return total_all_euros



    def total_euros():
        global total_extra_size_euros_label
        global total_a_size_euros_label
        global total_b_size_euros_label
        global total_c_size_euros_label
        global total_torn_euros_label
        global total_all_euros_label

        euros_w = Tk()
        euros_w.title("ΣΥΝΟΛΙΚΑ ΕΥΡΩ")
        euros_w.geometry("800x600")
        euros_w.config(background="#EEE8D1")


        total_extra_size_euros_btn = Button(euros_w, text="ΣΥΝΟΛΙΚΑ ΕΥΡΩ 'EXTRA'", command=total_extra_size_euros, font=("arial black", 11), bd=4, background="#DE9C30", activeforeground="#DE9C30", activebackground="#DE9C30")
        total_extra_size_euros_btn.grid(row=1, column=0, pady=(25, 5), padx=10, ipadx=32)
        total_extra_size_euros_label = Label(euros_w, text="    ", font=("arial black", 14), background="#A2D8F0",  relief="ridge", borderwidth=5, foreground="#42047A", width=15)
        total_extra_size_euros_label.grid(row=1, column=1, ipadx=100, pady=(25, 5))
        total_a_size_euros_btn = Button(euros_w, text="ΣΥΝΟΛΙΚΑ ΕΥΡΩ 'Α'", command=total_a_size_euros, font=("arial black", 11), bd=4, background="#DE9C30", activeforeground="#DE9C30", activebackground="#DE9C30")
        total_a_size_euros_btn.grid(row=2, column=0, pady=5, padx=10, ipadx=54)
        total_a_size_euros_label = Label(euros_w, text="    ", font=("arial black", 14), background="#A2D8F0",  relief="ridge", borderwidth=5, foreground="#42047A", width=15)
        total_a_size_euros_label.grid(row=2, column=1, ipadx=100, pady=5)
        total_b_size_euros_btn = Button(euros_w, text="ΣΥΝΟΛΙΚΑ ΕΥΡΩ 'Β'", command=total_b_size_euros, font=("arial black", 11), bd=4, background="#DE9C30", activeforeground="#DE9C30", activebackground="#DE9C30")
        total_b_size_euros_btn.grid(row=3, column=0, pady=5, padx=10, ipadx=54)
        total_b_size_euros_label = Label(euros_w, text="    ", font=("arial black", 14), background="#A2D8F0",  relief="ridge", borderwidth=5, foreground="#42047A", width=15)
        total_b_size_euros_label.grid(row=3, column=1, ipadx=100, pady=5)
        total_c_size_euros_btn = Button(euros_w, text="ΣΥΝΟΛΙΚΑ ΕΥΡΩ 'Γ'", command=total_c_size_euros, font=("arial black", 11), bd=4, background="#DE9C30", activeforeground="#DE9C30", activebackground="#DE9C30")
        total_c_size_euros_btn.grid(row=4, column=0, pady=5, padx=10, ipadx=54)
        total_c_size_euros_label = Label(euros_w, text="    ", font=("arial black", 14), background="#A2D8F0",  relief="ridge", borderwidth=5, foreground="#42047A", width=15)
        total_c_size_euros_label.grid(row=4, column=1, ipadx=100, pady=5)
        total_torn_euros_btn = Button(euros_w, text="ΣΥΝΟΛΙΚΑ ΕΥΡΩ 'ΣΚΙΣΜΕΝΑ'", command=total_torn_euros, font=("arial black", 11), bd=4, background="#DE9C30", activeforeground="#DE9C30", activebackground="#DE9C30")
        total_torn_euros_btn.grid(row=5, column=0, pady=5, padx=10, ipadx=15)
        total_torn_euros_label = Label(euros_w, text="    ", font=("arial black", 14), background="#A2D8F0",  relief="ridge", borderwidth=5, foreground="#42047A", width=15)
        total_torn_euros_label.grid(row=5, column=1, ipadx=100, pady=5)
        total_all_euros_btn = Button(euros_w, text="ΣΥΝΟΛΟ ΕΥΡΩ ", command=total_all_euros, font=("arial black", 13), bd=8, background="#1CC405", activeforeground="#1CC405", activebackground="#1CC405")
        total_all_euros_btn.grid(row=6, column=0, columnspan=2,  pady=40, padx=10, ipadx=80)
        total_all_euros_label = Label(euros_w, text="    ",font=("arial black", 15), background="#D1EFCD",
                             relief="ridge", borderwidth=10, foreground="#42047A", width=20)
        total_all_euros_label.grid(row=7, column=0, columnspan=2, ipadx=100, pady=15)

        title_label = Label(euros_w, text=" ΚΑΣΤΑΝΑ DATABASE ", font=("arial black", 18), background="black",
                            relief="ridge", borderwidth=10, foreground="red")
        title_label.grid(row=0, pady=10, padx=10, columnspan=2)




    def total_profit():
        profit_window = Tk()
        profit_window.title("ΣΥΝΟΛΙΚΟ ΚΕΡΔΟΣ")
        profit_window.geometry("600x450")
        profit_window.config(background="#EEE8D1")

        workers = Entry(profit_window, width=15, font=("arial black", 14), bd=2)
        workers.grid(row=1, column=1, padx=20, pady=(20, 5))
        pesticide = Entry(profit_window, width=15, font=("arial black", 14), bd=2)
        pesticide.grid(row=2, column=1, padx=20, pady=5)
        total_euros = Entry(profit_window, width=15, font=("arial black", 14), bd=2)
        total_euros.grid(row=3, column=1, padx=20, pady=5)

        workers_label = Label(profit_window, text="ΚΟΣΤΟΣ ΕΡΓΑΤΩΝ :", font=("arial black", 12),  background="#DE9C30",  relief="ridge", borderwidth=5, foreground="black", width=28)
        workers_label.grid(row=1, column=0, pady=(20, 0), sticky=W, padx=10)
        pesticide_label = Label(profit_window, text="ΦΥΤΟΦΑΡΜΑΚΑ & ΛΙΠΑΣΜΑΤΑ:", font=("arial black", 12),  background="#DE9C30",  relief="ridge", borderwidth=5, foreground="black", width=28)
        pesticide_label.grid(row=2, column=0, sticky=W, padx=10)
        total_euros_label = Label(profit_window, text="ΣΥΝΟΛΙΚΑ ΕΥΡΩ :", font=("arial black", 12),  background="#DE9C30",  relief="ridge", borderwidth=5, foreground="black", width=28)
        total_euros_label.grid(row=3, column=0, sticky=W, padx=10)

        title_label = Label(profit_window, text=" ΚΑΣΤΑΝΑ DATABASE ", font=("arial black", 18), background="black",
                            relief="ridge", borderwidth=10, foreground="red")
        title_label.grid(row=0, pady=10, padx=10, columnspan=2)



        def total_profit1():
            global total_all_euros_label

            workers_value = float(workers.get().replace(",", "."))
            pesticide_value = float(pesticide.get().replace(",", "."))
            total_euros_value = float(total_euros.get().replace(",", "."))



            total_profit = total_euros_value - workers_value - pesticide_value

            total_profit1_label.config(text="ΣΥΝΟΛΙΚΟ ΚΕΡΔΟΣ : {:.2f} €".format(total_profit))


        total_profit1_btn = Button(profit_window, text="ΣΥΝΟΛΙΚΟ ΚΕΡΔΟΣ:", font=("arial black", 13),
                                   command=total_profit1, bd=7,
                                   background="#1CC405", activebackground="#1CC405", activeforeground="#1CC405")

        total_profit1_btn.grid(row=4, padx=5, columnspan=2, pady=(50, 20), ipadx=25)

        total_profit1_label = Label(profit_window, text="ΣΥΝΟΛΙΚΟ ΚΕΡΔΟΣ  :  0 € ", font=("arial black", 14), width=30,
                                    background="#D1EFCD", relief="ridge", borderwidth=10, foreground="#42047A")
        total_profit1_label.grid(row=5, columnspan=2, padx=10, pady=10, ipadx=15)




    date_label = Label(root, text="  ΗΜΕΡ/ΝΙΑ :  ", font=("arial black", 13), background="#EEE8D1", relief="ridge", borderwidth=3, foreground="#03647A")
    date_label.grid(row=2, column=0, pady=(20, 0),sticky=W, padx=20)
    kilos_label = Label(root, text="ΚΙΛΑ", font=("arial black", 15), background="#8DDFE1", foreground="#8B0669", relief="ridge", bd=5, width=5)
    kilos_label.grid(row=2, column=1, sticky=W, padx=40, pady=(10, 0))
    price_label = Label(root, text="ΤΙΜΗ", font=("arial black", 15), background="#8DDFE1", foreground="#8B0669", relief="ridge", bd=5, width=5)
    price_label.grid(row=2, column=2, sticky=W, padx=40, pady=(10, 0))

    extra_size_kilos_label = Label(root, text="ΜΕΓΕΘΟΣ 'EXTRA' : ", font=("arial black", 12), foreground="#5D046B", background="#EEE8D1")
    extra_size_kilos_label.grid(row=3, column=0, sticky=W, padx=40, pady=7)
    a_size_kilos_label = Label(root, text="ΜΕΓΕΘΟΣ  'Α' : ", font=("arial black", 12), foreground="#5D046B", background="#EEE8D1")
    a_size_kilos_label.grid(row=4, column=0, sticky=W, padx=40, pady=7)
    b_size_kilos_label = Label(root, text="ΜΕΓΕΘΟΣ  'Β' : ", font=("arial black", 12), foreground="#5D046B", background="#EEE8D1")
    b_size_kilos_label.grid(row=5, column=0, sticky=W, padx=40, pady=7)
    c_size_kilos_label = Label(root, text="ΜΕΓΕΘΟΣ  'Γ' : ", font=("arial black", 12), foreground="#5D046B", background="#EEE8D1")
    c_size_kilos_label.grid(row=6, column=0, sticky=W, padx=40, pady=7)
    torn_kilos_label = Label(root, text="'ΣΚΙΣΜΕΝΑ' :", font=("arial black", 12), foreground="#5D046B", background="#EEE8D1")
    torn_kilos_label.grid(row=7, column=0, sticky=W, padx=40, pady=7)
    delete_box_label = Label(root, text="ΕΠΙΛΟΓΗ ID : ", font=("arial black", 13), relief="ridge", borderwidth=3, foreground="#03647A", background="#EEE8D1")
    delete_box_label.grid(row=8, column=0, sticky=W, padx=20, pady=10)

    date = Entry(root, width=10, font=("arial black", 13), background="#D7EEEE", bd=3)
    date.grid(row=2, column=0, columnspan=2, padx=30, pady=(20, 5))
    extra_size_kilos = Entry(root, width=12, font=("arial black", 12), background="#D6E3E1", bd=2)
    extra_size_kilos.grid(row=3, column=1, sticky=W, padx=10, pady=7)
    extra_size_price = Entry(root, width=12, font=("arial black", 12), background="#D6E3E1", bd=2)
    extra_size_price.grid(row=3, column=2, padx=10, sticky=W, pady=7)
    a_size_kilos = Entry(root, width=12, font=("arial black", 12), background="#D6E3E1", bd=2)
    a_size_kilos.grid(row=4, column=1, padx=10, sticky=W, pady=7)
    a_size_price = Entry(root, width=12, font=("arial black", 12), background="#D6E3E1", bd=2)
    a_size_price.grid(row=4, column=2, padx=10,  sticky=W, pady=7)
    b_size_kilos = Entry(root, width=12, font=("arial black", 12), background="#D6E3E1", bd=2)
    b_size_kilos.grid(row=5, column=1, padx=10, sticky=W, pady=7)
    b_size_price = Entry(root, width=12, font=("arial black", 12), background="#D6E3E1", bd=2)
    b_size_price.grid(row=5, column=2, padx=10,  sticky=W, pady=7)
    c_size_kilos = Entry(root, width=12, font=("arial black", 12), background="#D6E3E1", bd=2)
    c_size_kilos.grid(row=6, column=1, padx=10, sticky=W, pady=7)
    c_size_price = Entry(root, width=12, font=("arial black", 12), background="#D6E3E1", bd=2)
    c_size_price.grid(row=6, column=2, padx=10, sticky=W, pady=7)
    torn_kilos = Entry(root, width=12, font=("arial black", 12), background="#D6E3E1", bd=2)
    torn_kilos.grid(row=7, column=1, padx=10, sticky=W, pady=7)
    torn_price = Entry(root, width=12, font=("arial black", 12), background="#D6E3E1", bd=2)
    torn_price.grid(row=7, column=2, padx=10, sticky=W, pady=7)

    delete_box = Entry(root, width=10, font=("arial black", 14), background="#D7EEEE", bd=3)
    delete_box.grid(row=8, column=0, columnspan=2, padx=50, pady=10)



    submit_btn = Button(root, text="Εισαγωγή Στοιχείων", command=submit, font=("arial black", 12), bd=4, background="#06D784", activebackground="#06D784", activeforeground="#06D784")
    submit_btn.grid(row=9, column=0, pady=(20, 10), padx=10, ipadx=40)

    show_btn = Button(root, text="Εμφάνιση Στοιχείων", command=show, font=("arial black", 12), bd=4,  background="#0499A0", activebackground="#0499A0", activeforeground="#0499A0")
    show_btn.grid(row=10, column=1, columnspan=2, pady=10, padx=20, ipadx=40)

    delete_btn = Button(root, text="Διαγραφή Στοιχείων", command=delete, font=("arial black", 12), bd=4, background="#F67EA3", activebackground="#F67EA3", activeforeground="#F67EA3")
    delete_btn.grid(row=10, column=0, pady=10, padx=20, ipadx=40)

    edit_btn = Button(root, text="Επεξεργασία Στοιχείων", command=edit, font=("arial black", 12), bd=4, background="#A481C9", activebackground="#A481C9", activeforeground="#A481C9")
    edit_btn.grid(row=9, column=1, columnspan=2, pady=(20, 10), padx=20, ipadx=28)


    total_kilos_btn = Button(root, text="ΣΥΝΟΛΙΚΑ ΚΙΛΑ", command=open_kilos_w,  width=23, font=("arial black", 12), bd=7, background="#8CD0EC", activebackground="#A481C9", activeforeground="#A481C9")
    total_kilos_btn.grid(row=11, column=0, pady=(20, 10))


    total_euros_btn = Button(root, text="ΣΥΝΟΛΙΚΑ ΕΥΡΩ", command=total_euros, width=23, font=("arial black", 12), bd=7, background="#8CD0EC", activebackground="#A481C9", activeforeground="#A481C9")
    total_euros_btn.grid(row=11, column=1, columnspan=2, pady=(20, 10))

    average_price_btn = Button(root, text="ΜΕΣΗ ΤΙΜΗ", command=average_price, width=23, font=("arial black", 12), bd=7, background="#8CD0EC", activebackground="#A481C9", activeforeground="#A481C9")
    average_price_btn.grid(row=12, column=0, pady=(5, 10))

    total_profit_btn = Button(root, text="ΣΥΝΟΛΙΚΟ ΚΕΡΔΟΣ", command=total_profit, width=23, font=("arial black", 12), bd=7, background="#8CD0EC", activebackground="#A481C9", activeforeground="#A481C9")
    total_profit_btn.grid(row=12, column=1, columnspan=2, pady=(5, 5))


    name_label = Label(root, text="Created and Designed by : Papaioannou Antonios", font=("arial black", 10), foreground="grey", background="#EEE8D1", borderwidth=1)
    name_label.grid(column=1, row=13, columnspan=2, sticky=W, pady=(1, 0))






    conn.commit()
    conn.close()

    root.mainloop()